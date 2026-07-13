#!/usr/bin/env python3
"""
generate_production_log.py — Gopuff litigation production log (Excel)

Usage:
    python generate_production_log.py --index bates_index.json --output <output_dir> [--log-data log_data.json]

Reads bates_index.json (output of bates_stamp.py) and an optional log_data.json
with summaries and methods, then writes:
  - <PREFIX>_Production_Log.xlsx   (formatted Excel spreadsheet)

log_data.json format (optional — summaries and methods keyed by filename):
{
  "entries": [
    {
      "filename": "offer_letter.pdf",
      "summary": "Employment offer letter dated 3/15/2023 to plaintiff John Smith.",
      "method": "Electronic"
    },
    {
      "filename": "policy_ack.xlsx",
      "summary": "Employee handbook acknowledgment form signed by plaintiff.",
      "method": "Native"
    }
  ]
}

If log_data.json is not provided, summary will be blank and method will be
inferred from the native flag in bates_index.json.
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")


# ── Gopuff brand colors ───────────────────────────────────────────────────────

NAVY        = "0A1F44"
RED         = "E63329"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F4F7"
MID_GRAY    = "D0D5DD"
DARK_TEXT   = "1A1A2E"

# ── Helpers ───────────────────────────────────────────────────────────────────

def doc_type_from_filename(filename: str) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    TYPE_MAP = {
        "pdf":  "PDF",
        "doc":  "Word Document",
        "docx": "Word Document",
        "xls":  "Excel Spreadsheet",
        "xlsx": "Excel Spreadsheet",
        "xlsm": "Excel Spreadsheet",
        "ppt":  "PowerPoint",
        "pptx": "PowerPoint",
        "mp4":  "Video",
        "mov":  "Video",
        "avi":  "Video",
        "wmv":  "Video",
        "msg":  "Email (MSG)",
        "eml":  "Email (EML)",
        "jpg":  "Image",
        "jpeg": "Image",
        "png":  "Image",
        "tiff": "Image",
        "tif":  "Image",
        "gif":  "Image",
        "zip":  "Archive",
        "rar":  "Archive",
        "txt":  "Text File",
        "csv":  "CSV / Data File",
    }
    return TYPE_MAP.get(ext, ext.upper() if ext else "Unknown")


def infer_method(entry: dict) -> str:
    return "Native" if entry.get("native") else "Electronic"


def thin_border():
    side = Side(style="thin", color=MID_GRAY)
    return Border(top=side, bottom=side, left=side, right=side)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Gopuff production log generator")
    parser.add_argument("--index",    required=True, help="Path to bates_index.json")
    parser.add_argument("--output",   required=True, help="Output directory")
    parser.add_argument("--log-data", default=None,  help="Optional path to log_data.json (summaries + methods)")
    args = parser.parse_args()

    with open(args.index) as f:
        index = json.load(f)

    prefix    = index.get("prefix", "PROD")
    case_name = index.get("case_name", "")
    entries   = index.get("entries", [])

    log_data_map = {}
    if args.log_data and os.path.exists(args.log_data):
        with open(args.log_data) as f:
            log_data = json.load(f)
        for item in log_data.get("entries", []):
            log_data_map[item.get("filename", "")] = item

    os.makedirs(args.output, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Log"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:3"

    # Title row
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"DOCUMENT PRODUCTION LOG  |  {case_name}"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Subtitle row
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (
        f"Bates Range: {index.get('range_start','')}–{index.get('range_end','')}  |  "
        f"Total Documents: {len(entries)}  |  Total Pages: {index.get('total_pages', 0)}"
    )
    s.font = Font(name="Calibri", size=10, color=WHITE)
    s.fill = PatternFill("solid", fgColor=RED)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # Spacer
    ws.row_dimensions[3].height = 6

    # Header row
    HEADERS = ["#", "Document Title", "Document Type", "Summary",
               "Bates Start", "Bates End", "Pages", "Method of Production"]
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[4].height = 22

    # Data rows
    for i, entry in enumerate(entries):
        row = 5 + i
        filename = entry.get("filename", "")
        extra    = log_data_map.get(filename, {})

        values = [
            i + 1,
            entry.get("title", ""),
            extra.get("doc_type") or doc_type_from_filename(filename),
            extra.get("summary", ""),
            entry.get("bates_start", ""),
            entry.get("bates_end", ""),
            entry.get("pages", 0),
            extra.get("method") or infer_method(entry),
        ]

        row_fill = PatternFill("solid", fgColor=(LIGHT_GRAY if i % 2 == 0 else WHITE))
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = thin_border()

            if col_idx in (5, 6):  # Bates numbers
                cell.font = Font(name="Courier New", size=9, color=DARK_TEXT)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 1:
                cell.font = Font(name="Calibri", size=10, color=DARK_TEXT)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 7:
                cell.font = Font(name="Calibri", size=10, color=DARK_TEXT)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 4:
                cell.font = Font(name="Calibri", size=10, color=DARK_TEXT)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=10, color=DARK_TEXT)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.row_dimensions[row].height = 42

    # Totals row
    total_row = 5 + len(entries)
    ws.merge_cells(f"A{total_row}:F{total_row}")
    tl = ws.cell(row=total_row, column=1,
                 value=f"TOTAL  ({len(entries)} documents)")
    tl.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    tl.fill = PatternFill("solid", fgColor=NAVY)
    tl.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    tl.border = thin_border()

    tp = ws.cell(row=total_row, column=7, value=index.get("total_pages", 0))
    tp.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    tp.fill = PatternFill("solid", fgColor=NAVY)
    tp.alignment = Alignment(horizontal="center", vertical="center")
    tp.border = thin_border()

    me = ws.cell(row=total_row, column=8, value="")
    me.fill = PatternFill("solid", fgColor=NAVY)
    me.border = thin_border()
    ws.row_dimensions[total_row].height = 20

    # Column widths
    for col_idx, width in {1: 5, 2: 32, 3: 18, 4: 45, 5: 15, 6: 15, 7: 7, 8: 22}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"

    out_path = os.path.join(args.output, f"{prefix}_Production_Log.xlsx")
    wb.save(out_path)

    print(f"\n✓ Production log written: {out_path}")
    print(f"  Documents: {len(entries)}")
    print(f"  Total pages: {index.get('total_pages', 0)}")
    print()


if __name__ == "__main__":
    main()
